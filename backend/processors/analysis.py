from __future__ import annotations

import asyncio
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional

from ..analysis_components import AnalysisEngine, PDFDocumentLoader, build_default_engine
from ..document_analysis import (
    AzureDocumentIntelligenceExtractor,
    FormatGuidedExtractor,
    FormatRepository,
    LabelAnalysisService,
)

from backend.jobs.models import JobCompletion, JobRecord
from backend.jobs.worker import JobProcessor, ProgressReporter

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class AnalysisRow:
    id: int
    filename: str
    model_name: str
    voltage: str
    typ_batt_capacity_wh: str
    typ_capacity_mah: str
    rated_capacity_mah: str
    rated_energy_wh: str

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "filename": self.filename,
            "model_name": self.model_name,
            "voltage": self.voltage,
            "typ_batt_capacity_wh": self.typ_batt_capacity_wh,
            "typ_capacity_mah": self.typ_capacity_mah,
            "rated_capacity_mah": self.rated_capacity_mah,
            "rated_energy_wh": self.rated_energy_wh,
        }


class AnalysisJobProcessor(JobProcessor):
    def __init__(
        self,
        *,
        document_loader: Optional[PDFDocumentLoader] = None,
        analysis_engine: Optional[AnalysisEngine] = None,
        label_service_factory: Optional[Callable[[], LabelAnalysisService]] = None,
        format_repository_dir: Optional[str | Path] = None,
        sleep_seconds: float = 0.05,
    ) -> None:
        self._document_loader = document_loader or PDFDocumentLoader()
        self._analysis_engine = analysis_engine or build_default_engine()
        self._format_repository_dir = (
            Path(format_repository_dir) if format_repository_dir is not None else None
        )
        self._label_service_factory = label_service_factory
        self._sleep_seconds = sleep_seconds

    async def run(self, job: JobRecord, job_dir: Path, reporter: ProgressReporter) -> JobCompletion:
        label_service = self._initialise_label_service()
        try:
            return await self._execute(job, job_dir, reporter, label_service)
        finally:
            await self._close_label_service(label_service)

    async def _execute(
        self,
        job: JobRecord,
        job_dir: Path,
        reporter: ProgressReporter,
        label_service: LabelAnalysisService,
    ) -> JobCompletion:
        input_dir = job_dir / "input"
        output_dir = job_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        files = job.input_manifest
        total = len(files)
        if total == 0:
            await reporter.report(
                processed=0,
                total=0,
                current_file=None,
                message="No PDF files selected for analysis.",
            )
            return JobCompletion(output_manifest=[], download_path=None)

        await reporter.report(
            processed=0,
            total=total,
            current_file=None,
            message=f"Queued {total} file(s) for processing.",
        )

        results: List[AnalysisRow] = []
        for index, entry in enumerate(files, start=1):
            filename = entry.get("filename")
            source_path = input_dir / filename
            if not source_path.exists():
                raise FileNotFoundError(f"Input file not found: {source_path}")

            await reporter.report(
                processed=index - 1,
                total=total,
                current_file=filename,
                message=f"Processing {filename}",
            )

            fields, messages = await label_service.analyse(source_path)
            for message in messages:
                await reporter.report(
                    processed=index - 1,
                    total=total,
                    current_file=filename,
                    message=message,
                )

            results.append(self._build_row(index, filename, fields))

            await reporter.report(
                processed=index,
                total=total,
                current_file=filename,
                message=f"Completed {filename}",
            )

            if self._sleep_seconds:
                await asyncio.sleep(self._sleep_seconds)

        if not results:
            await reporter.report(
                processed=total,
                total=total,
                current_file=None,
                message="No data extracted from the provided PDFs.",
            )
            return JobCompletion(output_manifest=[], download_path=None)

        report_path = self._write_excel(output_dir, results)
        await reporter.report(
            processed=total,
            total=total,
            current_file=None,
            message=f"Exported results to {report_path.name}",
        )

        manifest = [row.to_dict() for row in results]
        return JobCompletion(output_manifest=manifest, download_path=str(report_path))

    def _initialise_label_service(self) -> LabelAnalysisService:
        if self._label_service_factory:
            return self._label_service_factory()

        repository: Optional[FormatRepository] = None
        extractor: Optional[FormatGuidedExtractor] = None
        di_extractor: Optional[AzureDocumentIntelligenceExtractor] = None

        target_dir = self._format_repository_dir
        if target_dir is None:
            env_dir = os.getenv("ANALYSIS_FORMAT_DIR")
            if env_dir:
                target_dir = Path(env_dir)
            else:
                target_dir = Path(__file__).resolve().parent.parent / "formats"

        if target_dir and target_dir.exists():
            repository = FormatRepository(target_dir)
            extractor = FormatGuidedExtractor()

        try:
            di_extractor = AzureDocumentIntelligenceExtractor()
        except Exception:  # pragma: no cover - optional dependency
            di_extractor = None

        return LabelAnalysisService(
            document_loader=self._document_loader,
            analysis_engine=self._analysis_engine,
            format_repository=repository,
            extractor=extractor,
            document_intelligence_extractor=di_extractor,
        )

    def _build_row(self, index: int, filename: str, fields: Dict[str, Any]) -> AnalysisRow:
        def as_text(key: str) -> str:
            value = fields.get(key, "")
            if value is None:
                return ""
            return str(value).strip()

        return AnalysisRow(
            id=index,
            filename=filename,
            model_name=as_text("model_name"),
            voltage=as_text("voltage"),
            typ_batt_capacity_wh=as_text("typ_batt_capacity_wh"),
            typ_capacity_mah=as_text("typ_capacity_mah"),
            rated_capacity_mah=as_text("rated_capacity_mah"),
            rated_energy_wh=as_text("rated_energy_wh"),
        )

    def _write_excel(self, output_dir: Path, rows: Iterable[AnalysisRow]) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
        except ImportError as exc:  # pragma: no cover - optional dependency check
            raise RuntimeError("openpyxl is required to export Excel reports") from exc

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Analysis"
        headers = [
            "ID",
            "Filename",
            "Model Name",
            "Voltage",
            "Typ Batt Capacity Wh",
            "Typ Capacity mAh",
            "Rated Capacity mAh",
            "Rated Energy Wh",
        ]
        worksheet.append(headers)

        highlight = PatternFill(start_color="FFFDEB95", end_color="FFFDEB95", fill_type="solid")
        for row_index, row in enumerate(rows, start=2):
            values = [
                row.id,
                row.filename,
                row.model_name,
                row.voltage,
                row.typ_batt_capacity_wh,
                row.typ_capacity_mah,
                row.rated_capacity_mah,
                row.rated_energy_wh,
            ]
            worksheet.append(values)
            for column, value in enumerate(values, start=1):
                if column >= 3 and (value is None or str(value).strip() == ""):
                    worksheet.cell(row=row_index, column=column).fill = highlight

        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / "analysis_result.xlsx"
        workbook.save(path)
        return path

    async def _close_label_service(self, service: LabelAnalysisService) -> None:
        closer = getattr(service, "aclose", None)
        if closer is not None:
            await closer()
