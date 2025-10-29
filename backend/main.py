import asyncio
import os
import logging # Import the logging module
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from uuid import uuid4

from fastapi import Depends, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from analysis_components import (
    AnalysisEngine,
    PDFDocumentLoader,
    build_default_engine,
)
from document_analysis import (
    AzureDocumentIntelligenceExtractor,
    FormatGuidedExtractor,
    FormatRepository,
    LabelAnalysisService,
)
from settings import ensure_env_loaded

ensure_env_loaded()

# Get a logger instance for this module
logger = logging.getLogger(__name__)
logging.getLogger('azure.core.pipeline.policies.http_logging_policy').setLevel(logging.WARNING)
logging.getLogger('httpx').setLevel(logging.WARNING)
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover - defensive guard for runtime
    Workbook = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]

BASE_WORKDIR = Path(
    os.getenv("ANALYSIS_JOBS_DIR", Path(__file__).resolve().parent / "job_runs")
).resolve()
BASE_WORKDIR.mkdir(parents=True, exist_ok=True)
DEFAULT_MAX_CONCURRENT_JOBS = max(1, int(os.getenv("ANALYSIS_MAX_CONCURRENT", "2")))


class PDFFile(BaseModel):
    id: int
    filename: str


class ListPDFsRequest(BaseModel):
    path: str


class AnalyzeRequest(BaseModel):
    path: str
    files: List[PDFFile]


class AnalysisResult(BaseModel):
    id: int
    filename: str
    model_name: str
    voltage: str
    typ_batt_capacity_wh: str
    typ_capacity_mah: str
    rated_capacity_mah: str
    rated_energy_wh: str


class JobStatus(str, Enum):
    QUEUED = "queued"
    RUNNING = "running"
    COMPLETED = "completed"
    CANCELLED = "cancelled"
    FAILED = "failed"


class StartAnalysisResponse(BaseModel):
    job_id: str
    status: JobStatus


class AnalysisStatusResponse(BaseModel):
    job_id: str
    status: JobStatus
    progress: float
    processed_count: int
    total_count: int
    results: List[AnalysisResult]
    download_ready: bool
    download_path: str | None
    error: str | None = None
    current_file: str | None = None
    messages: List[str] = []


class StopAnalysisResponse(AnalysisStatusResponse):
    pass


@dataclass
class PipelineResult:
    download_path: Optional[Path]
    error: Optional[str] = None


@dataclass
class AnalysisJobState:
    job_id: str
    request: AnalyzeRequest
    job_dir: Path
    status: JobStatus = JobStatus.QUEUED
    progress: float = 0.0
    processed_count: int = 0
    total_count: int = 0
    results: List[AnalysisResult] = field(default_factory=list)
    current_file: Optional[str] = None
    download_path: Optional[Path] = None
    error: Optional[str] = None
    messages: List[str] = field(default_factory=list)
    task: Optional[asyncio.Task] = None
    cancel_event: asyncio.Event = field(default_factory=asyncio.Event)
    lock: asyncio.Lock = field(default_factory=asyncio.Lock)

    async def snapshot(self) -> AnalysisStatusResponse:
        async with self.lock:
            download_ready = bool(self.download_path and self.download_path.exists())
            return AnalysisStatusResponse(
                job_id=self.job_id,
                status=self.status,
                progress=round(self.progress, 2),
                processed_count=self.processed_count,
                total_count=self.total_count,
                results=[result.model_copy(deep=True) for result in self.results],
                download_ready=download_ready,
                download_path=str(self.download_path) if download_ready else None,
                error=self.error,
                current_file=self.current_file,
                messages=list(self.messages),
            )

    async def update_total(self, total: int) -> None:
        async with self.lock:
            self.total_count = total

    async def add_message(self, message: str) -> None:
        async with self.lock:
            self.messages.append(message)
            logger.info(f"Job {self.job_id}: {message}") # Log the message


class JobCallbacks:
    def __init__(self, job: AnalysisJobState) -> None:
        self.job = job

    async def set_total(self, total_count: int) -> None:
        await self.job.update_total(total_count)

    async def set_current_file(self, filename: Optional[str]) -> None:
        async with self.job.lock:
            self.job.current_file = filename

    async def add_result(self, result: AnalysisResult) -> None:
        async with self.job.lock:
            self.job.results.append(result)
            self.job.processed_count = len(self.job.results)
            if self.job.total_count:
                self.job.progress = min(100.0, (self.job.processed_count / self.job.total_count) * 100)

    async def add_message(self, message: str) -> None:
        await self.job.add_message(message)

    def should_cancel(self) -> bool:
        return self.job.cancel_event.is_set()

    async def mark_download(self, output_path: Optional[Path]) -> None:
        async with self.job.lock:
            self.job.download_path = output_path
            if output_path and self.job.status == JobStatus.RUNNING:
                self.job.progress = 100.0


class AnalysisPipeline:
    async def run(self, job: AnalysisJobState, callbacks: JobCallbacks) -> PipelineResult:
        raise NotImplementedError


class DefaultAnalysisPipeline(AnalysisPipeline):
    def __init__(
        self,
        *,
        document_loader: Optional[PDFDocumentLoader] = None,
        analysis_engine: Optional[AnalysisEngine] = None,
        sleep_seconds: float = 0.05,
        label_service: Optional[LabelAnalysisService] = None,
        format_repository_dir: Optional[str | Path] = None,
    ) -> None:
        self.document_loader = document_loader or PDFDocumentLoader()
        self.analysis_engine = analysis_engine or build_default_engine()
        self.sleep_seconds = sleep_seconds
        self.label_service = label_service or self._build_label_service(format_repository_dir)

    def _build_label_service(self, format_repository_dir: Optional[str | Path]) -> LabelAnalysisService:
        repository: Optional[FormatRepository] = None
        extractor: Optional[FormatGuidedExtractor] = None
        di_extractor: Optional[AzureDocumentIntelligenceExtractor] = None

        target_dir: Optional[Path] = None
        if format_repository_dir is not None:
            target_dir = Path(format_repository_dir)
        else:
            env_dir = os.getenv("ANALYSIS_FORMAT_DIR")
            if env_dir:
                target_dir = Path(env_dir)
            else:
                target_dir = Path(__file__).resolve().parent / "formats"

        if target_dir and target_dir.exists():
            repository = FormatRepository(target_dir)
            extractor = FormatGuidedExtractor()

        try:
            di_extractor = AzureDocumentIntelligenceExtractor()
        except Exception:
            di_extractor = None

        return LabelAnalysisService(
            document_loader=self.document_loader,
            analysis_engine=self.analysis_engine,
            format_repository=repository,
            extractor=extractor,
            document_intelligence_extractor=di_extractor,
        )

    async def run(self, job: AnalysisJobState, callbacks: JobCallbacks) -> PipelineResult:
        try:
            return await self._run_internal(job, callbacks)
        finally:
            await self._close_label_service()

    async def _run_internal(self, job: AnalysisJobState, callbacks: JobCallbacks) -> PipelineResult:
        source_path = Path(job.request.path)
        if not source_path.exists():
            raise FileNotFoundError(f"�䤣��ӷ����|: {job.request.path}")

        files = job.request.files
        if not files:
            await callbacks.add_message("No PDF files to analyse.")
            await callbacks.mark_download(None)
            return PipelineResult(download_path=None)

        await callbacks.set_total(len(files))
        job.job_dir.mkdir(parents=True, exist_ok=True)

        results: List[AnalysisResult] = []

        for index, pdf_file in enumerate(files, start=1):
            if callbacks.should_cancel():
                await callbacks.add_message("Analysis cancelled by user.")
                await callbacks.set_current_file(None)
                return PipelineResult(download_path=None)

            file_path = source_path / pdf_file.filename
            if not file_path.exists():
                raise FileNotFoundError(f"�䤣���ɮ�: {file_path}")

            await callbacks.set_current_file(pdf_file.filename)
            await callbacks.add_message(f"Starting analysis {pdf_file.filename}")

            try:
                raw_fields, extra_messages = await self._analyse_document(file_path)
            except Exception as exc:
                await callbacks.add_message(f"{pdf_file.filename} analysis failed: {exc}")
                raise

            for message in extra_messages:
                await callbacks.add_message(message)

            result = self._build_result(index, pdf_file.filename, raw_fields)
            await callbacks.add_result(result)

            logger.info(
                "%s fields: model_name=%r, voltage=%r, typ_batt_capacity_wh=%r, "
                "typ_capacity_mah=%r, rated_capacity_mah=%r, rated_energy_wh=%r",
                pdf_file.filename,
                result.model_name,
                result.voltage,
                result.typ_batt_capacity_wh,
                result.typ_capacity_mah,
                result.rated_capacity_mah,
                result.rated_energy_wh,
            )

            results.append(result)
            await callbacks.add_message(f"{pdf_file.filename} completed ({index}/{len(files)})")

            if self.sleep_seconds:
                await asyncio.sleep(self.sleep_seconds)

        if not results:
            await callbacks.add_message("No analysis results generated.")
            await callbacks.mark_download(None)
            return PipelineResult(download_path=None)

        excel_path = self._write_excel(job.job_dir, results)
        await callbacks.add_message(f"Exported analysis results to {excel_path.name}")
        await callbacks.mark_download(excel_path)
        await callbacks.set_current_file(None)
        return PipelineResult(download_path=excel_path)

    async def _close_label_service(self) -> None:
        closer = getattr(self.label_service, "aclose", None)
        if closer is None:
            return
        await closer()

    def _build_result(self, index: int, filename: str, fields: Dict[str, Any]) -> AnalysisResult:
        def as_text(key: str) -> str:
            value = fields.get(key, '')
            if value is None:
                return ''
            return str(value).strip()

        return AnalysisResult(
            id=index,
            filename=filename,
            model_name=as_text('model_name'),
            voltage=as_text('voltage'),
            typ_batt_capacity_wh=as_text('typ_batt_capacity_wh'),
            typ_capacity_mah=as_text('typ_capacity_mah'),
            rated_capacity_mah=as_text('rated_capacity_mah'),
            rated_energy_wh=as_text('rated_energy_wh'),
        )

    async def _analyse_document(self, file_path: Path) -> tuple[Dict[str, Any], List[str]]:
        if self.label_service:
            return await self.label_service.analyse(file_path)
        document = await asyncio.to_thread(self.document_loader.load, file_path)
        fields = await self.analysis_engine.analyse(document)
        return fields, []

    def _write_excel(self, job_dir: Path, results: List[AnalysisResult]) -> Path:
        if Workbook is None:
            raise RuntimeError('openpyxl 未安裝，無法匯出 Excel')
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Analysis'
        headers = [
            'ID',
            'Filename',
            'Model Name',
            'Voltage',
            'Typ Batt Capacity Wh',
            'Typ Capacity mAh',
            'Rated Capacity mAh',
            'Rated Energy Wh',
        ]
        worksheet.append(headers)
        missing_fill = None
        if PatternFill is not None:
            missing_fill = PatternFill(start_color="FFFDEB95", end_color="FFFDEB95", fill_type="solid")
        for index, item in enumerate(results, start=2):
            row_values = [
                item.id,
                item.filename,
                item.model_name,
                item.voltage,
                item.typ_batt_capacity_wh,
                item.typ_capacity_mah,
                item.rated_capacity_mah,
                item.rated_energy_wh,
            ]
            for column, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=index, column=column, value=value)
                if (
                    missing_fill is not None
                    and column >= 3
                    and (value is None or str(value).strip() == "")
                ):
                    cell.fill = missing_fill
        job_dir.mkdir(parents=True, exist_ok=True)
        output_path = job_dir / 'analysis_result.xlsx'
        workbook.save(output_path)
        return output_path

class AnalysisManager:
    def __init__(
        self,
        pipeline_factory: Callable[[], AnalysisPipeline],
        *,
        max_concurrent_jobs: int | None = None,
        base_dir: Optional[Path] = None,
    ) -> None:
        self.pipeline_factory = pipeline_factory
        self.jobs: Dict[str, AnalysisJobState] = {}
        self.lock = asyncio.Lock()
        self.max_concurrent_jobs = max(
            1, max_concurrent_jobs if max_concurrent_jobs is not None else DEFAULT_MAX_CONCURRENT_JOBS
        )
        self.base_dir = Path(base_dir) if base_dir is not None else BASE_WORKDIR
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self._semaphore = asyncio.Semaphore(self.max_concurrent_jobs)

    async def start_job(self, request: AnalyzeRequest | Dict[str, Any]) -> StartAnalysisResponse:
        payload = request if isinstance(request, AnalyzeRequest) else AnalyzeRequest(**request)
        job_id = uuid4().hex
        job_dir = self.base_dir / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        job = AnalysisJobState(job_id=job_id, request=payload, job_dir=job_dir)
        async with self.lock:
            self.jobs[job_id] = job

        await job.add_message("工作已加入佇列，等待開始。")
        job.task = asyncio.create_task(self._job_runner(job))
        return StartAnalysisResponse(job_id=job_id, status=job.status)

    async def get_status(self, job_id: str) -> AnalysisStatusResponse:
        job = await self._get_job(job_id)
        return await job.snapshot()

    async def stop_job(self, job_id: str) -> StopAnalysisResponse:
        job = await self._get_job(job_id)
        job.cancel_event.set()
        if job.task and not job.task.done():
            try:
                await job.task
            except asyncio.CancelledError:
                pass
        async with job.lock:
            if job.status == JobStatus.RUNNING:
                job.status = JobStatus.CANCELLED
                job.progress = min(job.progress, 99.0)
        return await job.snapshot()  # type: ignore[return-value]

    async def get_download_path(self, job_id: str) -> Optional[Path]:
        job = await self._get_job(job_id)
        async with job.lock:
            return job.download_path if job.download_path and job.download_path.exists() else None

    async def _run_job(self, job: AnalysisJobState) -> None:
        async with job.lock:
            job.status = JobStatus.RUNNING
            job.progress = 0.0
            job.messages.clear()

        pipeline = self.pipeline_factory()
        callbacks = JobCallbacks(job)
        try:
            result = await pipeline.run(job, callbacks)
            if job.cancel_event.is_set():
                async with job.lock:
                    job.status = JobStatus.CANCELLED
                    job.progress = min(job.progress, 99.0)
            elif result.error:
                async with job.lock:
                    job.status = JobStatus.FAILED
                    job.error = result.error
            else:
                async with job.lock:
                    job.status = JobStatus.COMPLETED
                    job.progress = 100.0
        except asyncio.CancelledError:
            async with job.lock:
                job.status = JobStatus.CANCELLED
        except Exception as exc:  # pragma: no cover - sentry for unexpected errors
            async with job.lock:
                job.status = JobStatus.FAILED
                job.error = str(exc)
                logger.exception(f"Job {job.job_id} failed with an unexpected error.") # Log the exception
        finally:
            async with job.lock:
                job.current_file = None

    async def _job_runner(self, job: AnalysisJobState) -> None:
        async with self._semaphore:
            if job.cancel_event.is_set():
                async with job.lock:
                    job.status = JobStatus.CANCELLED
                    job.progress = 0.0
                    job.current_file = None
                return
            await self._run_job(job)

    async def _get_job(self, job_id: str) -> AnalysisJobState:
        async with self.lock:
            if job_id not in self.jobs:
                raise HTTPException(status_code=404, detail=f"找不到工作 {job_id}")
            return self.jobs[job_id]


def get_analysis_manager() -> AnalysisManager:
    global analysis_manager
    return analysis_manager


app = FastAPI()

origins = [
    "http://localhost",
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
async def read_root() -> Dict[str, str]:
    return {"message": "Welcome to the ASUS Label Analysis Backend!"}


def _normalise_path(path_str: str) -> Path:
    if path_str.startswith("\\\\"):
        return Path(path_str)
    return Path(path_str)


@app.post("/api/list-pdfs", response_model=List[PDFFile])
async def list_pdfs(request: ListPDFsRequest) -> List[PDFFile]:
    target_path = _normalise_path(request.path)
    if not target_path.exists():
        raise HTTPException(status_code=404, detail=f"Path not found: {request.path}")
    if not target_path.is_dir():
        raise HTTPException(status_code=400, detail=f"Path is not a directory: {request.path}")

    pdf_files: List[PDFFile] = []
    try:
        entries = sorted(entry for entry in target_path.iterdir() if entry.is_file() and entry.suffix.lower() == ".pdf")
        for index, entry in enumerate(entries, start=1):
            pdf_files.append(PDFFile(id=index, filename=entry.name))
    except Exception as exc:  # pragma: no cover - filesystem guard
        raise HTTPException(status_code=500, detail=f"Error listing files: {exc}") from exc
    return pdf_files


@app.post("/api/analyze/start", response_model=StartAnalysisResponse)
async def start_analysis(
    request: AnalyzeRequest,
    manager: AnalysisManager = Depends(get_analysis_manager),
) -> StartAnalysisResponse:
    return await manager.start_job(request)


@app.get("/api/analyze/status/{job_id}", response_model=AnalysisStatusResponse)
async def get_analysis_status(
    job_id: str,
    manager: AnalysisManager = Depends(get_analysis_manager),
) -> AnalysisStatusResponse:
    return await manager.get_status(job_id)


@app.post("/api/analyze/stop/{job_id}", response_model=StopAnalysisResponse)
async def stop_analysis(
    job_id: str,
    manager: AnalysisManager = Depends(get_analysis_manager),
) -> StopAnalysisResponse:
    return await manager.stop_job(job_id)


@app.get("/api/analyze/download/{job_id}")
async def download_analysis(
    job_id: str,
    manager: AnalysisManager = Depends(get_analysis_manager),
) -> FileResponse:
    download_path = await manager.get_download_path(job_id)
    if not download_path:
        raise HTTPException(status_code=404, detail="分析結果尚未準備好。")
    return FileResponse(
        path=download_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_path.name,
    )


analysis_manager = AnalysisManager(pipeline_factory=lambda: DefaultAnalysisPipeline())


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
